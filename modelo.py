from openpyxl import load_workbook

def substituir_planilha(
    modelo_path,
    nome,
    disciplina,
    serie,
    turma,
    aulas_semanais,
    recursos,
    # Datas (início e fim de cada semana)
    datas_semanas,
    # Para cada semana: (aulas, objetivo, objeto, habilidades, metodologia, proposta)
    semanas_info,
    output_path="PlanoGerado.xlsx"
):
    """
    Substitui todos os placeholders na aba 'Agosto' do modelo.

    Argumentos:
        modelo_path: Caminho para o modelo Excel.
        nome, disciplina, serie, turma: Cabeçalho.
        aulas_semanais: Nº de aulas por semana (preenche o cabeçalho).
        recursos: Recursos/Plataforma (mesmo para todas as semanas).
        datas_semanas: Lista com 4 tuplas (inicio, fim) para S1 a S4.
        semanas_info: Lista com 4 dicionários, cada um contendo:
            {
                "aulas": str,
                "objetivo": str,
                "objeto": str,
                "habilidades": str,
                "metodologia": str,
                "proposta": str
            }
        output_path: Nome do arquivo gerado.
    """
    wb = load_workbook(modelo_path)
    if "Agosto" in wb.sheetnames:
        ws = wb["Agosto"]
    else:
        ws = wb.active

    # Prepara o mapa de substituição
    mapa = {
        "#Nome": nome,
        "#Disciplina": disciplina,
        "#Ano": serie,
        "#Turmas": turma,
        "#AulasSemana": aulas_semanais,
        "#Recursos": recursos
    }

    # Adiciona dados de cada semana (1 a 4)
    for i in range(4):
        prefix = f"S{i+1}"
        data_inicio, data_fim = datas_semanas[i]
        info = semanas_info[i]

        mapa.update({
            f"#Data_{prefix}_inicio": data_inicio,
            f"#Data_{prefix}_fim": data_fim,
            f"#Aulas_{prefix}": info.get("aulas", ""),
            f"#Objetivo_{prefix}": info.get("objetivo", ""),
            f"#Objeto_{prefix}": info.get("objeto", ""),
            f"#Habilidades_{prefix}": info.get("habilidades", ""),
            f"#Metodologia_{prefix}": info.get("metodologia", ""),
            f"#Proposta_{prefix}": info.get("proposta", "")
        })

    # Faz a substituição nas células
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for chave, valor in mapa.items():
                    if chave in cell.value:
                        cell.value = cell.value.replace(chave, str(valor))

    wb.save(output_path)
    return output_path


# Teste rápido
# if __name__ == "__main__":
#     # Datas de exemplo (4 semanas)
#     datas = [
#         ("05/08/2025", "09/08/2025"),
#         ("12/08/2025", "16/08/2025"),
#         ("19/08/2025", "23/08/2025"),
#         ("26/08/2025", "30/08/2025")
#     ]

#     # Dados por semana (exemplo)
#     semanas = []
#     for i in range(4):
#         semanas.append({
#             "aulas": f"Aulas {i*4+1} a {i*4+4}",
#             "objetivo": f"Objetivo geral da semana {i+1}.",
#             "objeto": f"Conteúdo/Objeto da semana {i+1}.",
#             "habilidades": f"EF09MA1{i+1}, EF09MA2{i+1}",
#             "metodologia": f"Metodologia planejada para semana {i+1}.",
#             "proposta": f"Proposta para alunos elegíveis na semana {i+1}."
#         })

#     resultado = substituir_planilha(
#         modelo_path="./MODELO PLANO DE AULA PARCIAIS.xlsx",
#         nome="Anderson",
#         disciplina="Matemática",
#         serie="3º EM",
#         turma="A",
#         aulas_semanais="4",
#         recursos="- Livro do Estudante\n- Plataforma Digital",
#         datas_semanas=datas,
#         semanas_info=semanas
#     )

#     print(f"Plano gerado: {resultado}")

